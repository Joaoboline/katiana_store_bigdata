import os
import pandas as pd
import plotly.graph_objects as go
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import warnings
import datetime

warnings.filterwarnings("ignore")


input_excel_path = "data/vendas_input.xlsx"
output_excel = "out/katiana_excel_plotly_dashboard.xlsx"
plots_path = "plots"

os.makedirs("out", exist_ok=True)
os.makedirs("plots", exist_ok=True)


raw = pd.read_excel(input_excel_path, sheet_name="Vendas")
raw.columns = [col.strip().lower().replace(" ", "_") for col in raw.columns]


if "preco" in raw.columns:
    raw["valor_total"] = raw["preco"]
else:
    raise ValueError("A coluna 'preco' não foi encontrada na planilha.")

MARGEM_LUCRO = 0.30
raw["Lucro (R$)"] = raw["valor_total"] * MARGEM_LUCRO
raw["Capital Investido (R$)"] = raw["valor_total"] - raw["Lucro (R$)"]

print("✅ Colunas 'valor_total', 'Lucro (R$)' e 'Capital Investido (R$)' criadas com sucesso.")


raw["data"] = pd.to_datetime(raw["data"], errors="coerce")
raw = raw.dropna(subset=["data"])
raw = raw.sort_values("data")


df_diario = raw.groupby("data")[["valor_total", "Lucro (R$)", "Capital Investido (R$)"]].sum().reset_index()


fig_receita = go.Figure()
fig_receita.add_trace(go.Scatter(
    x=df_diario["data"],
    y=df_diario["valor_total"],
    mode="lines+markers",
    name="Receita (R$)",
    line=dict(color="royalblue"),
))
fig_receita.update_layout(title="📈 Receita Diária — Katiana Store", xaxis_title="Data", yaxis_title="Receita Total (R$)")
fig_receita.write_html(os.path.join(plots_path, "plot_receita_diaria.html"))


fig_lucro = go.Figure()
fig_lucro.add_trace(go.Bar(
    x=df_diario["data"],
    y=df_diario["Lucro (R$)"],
    name="Lucro Diário (R$)",
    marker_color="green",
))
fig_lucro.update_layout(title="💵 Lucro Diário — Margem de 30%", xaxis_title="Data", yaxis_title="Lucro (R$)")
fig_lucro.write_html(os.path.join(plots_path, "plot_lucro_diario.html"))


fig_lucro_capital = go.Figure()
fig_lucro_capital.add_trace(go.Bar(
    x=df_diario["data"],
    y=df_diario["Capital Investido (R$)"],
    name="Capital Investido (R$)",
    marker_color="gray",
))
fig_lucro_capital.add_trace(go.Bar(
    x=df_diario["data"],
    y=df_diario["Lucro (R$)"],
    name="Lucro (R$)",
    marker_color="green",
))
fig_lucro_capital.add_trace(go.Scatter(
    x=df_diario["data"],
    y=df_diario["valor_total"],
    mode="lines+markers",
    name="Receita Total (R$)",
    line=dict(color="royalblue", width=2),
))
fig_lucro_capital.update_layout(
    barmode="group",
    title="💰 Lucro x Capital Investido x Receita",
    xaxis_title="Data",
    yaxis_title="Valor (R$)",
)
fig_lucro_capital.write_html(os.path.join(plots_path, "plot_lucro_capital.html"))

print("📊 Gráficos de receita, lucro e capital investido gerados com sucesso!")


df_diario = df_diario.set_index("data")
df_diario = df_diario.asfreq("D", fill_value=0)


df_diario["valor_total_suave"] = df_diario["valor_total"].rolling(window=3, min_periods=1).mean()


modelo = ExponentialSmoothing(df_diario["valor_total_suave"], trend="add", seasonal=None)
resultado = modelo.fit()
forecast = resultado.forecast(30)


forecast_df = pd.DataFrame({
    "data": pd.date_range(df_diario.index[-1] + pd.Timedelta(days=1), periods=30),
    "Previsão (R$)": forecast
})


media = df_diario["valor_total"].mean()
forecast_df["Limite Inferior (R$)"] = forecast_df["Previsão (R$)"].apply(lambda x: max(x * 0.85, 0))
forecast_df["Limite Superior (R$)"] = forecast_df["Previsão (R$)"].apply(lambda x: min(x * 1.15, media * 2))

forecast_df = forecast_df[["data", "Previsão (R$)", "Limite Inferior (R$)", "Limite Superior (R$)"]]


fig_forecast = go.Figure()
fig_forecast.add_trace(go.Scatter(
    x=df_diario.index,
    y=df_diario["valor_total"],
    mode="lines",
    name="Histórico de Receita",
    line=dict(color="royalblue"),
))
fig_forecast.add_trace(go.Scatter(
    x=forecast_df["data"],
    y=forecast_df["Previsão (R$)"],
    mode="lines+markers",
    name="Previsão de Receita",
    line=dict(color="orange"),
))
fig_forecast.add_trace(go.Scatter(
    x=list(forecast_df["data"]) + list(forecast_df["data"][::-1]),
    y=list(forecast_df["Limite Superior (R$)"]) + list(forecast_df["Limite Inferior (R$)"][::-1]),
    fill="toself", fillcolor="rgba(255,165,0,0.2)",
    line=dict(color="rgba(255,255,255,0)"), hoverinfo="skip", showlegend=True,
    name="Intervalo de Confiança (Ajustado)",
))
fig_forecast.update_layout(
    title="📊 Previsão de Receita — Próximos 30 Dias (Ajustado)",
    xaxis_title="Data",
    yaxis_title="Receita Total (R$)",
)
fig_forecast.write_html(os.path.join(plots_path, "plot_previsao_30_dias.html"))


if os.path.exists(output_excel):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_excel = f"out/katiana_excel_plotly_dashboard_{timestamp}.xlsx"

with pd.ExcelWriter(output_excel, engine="xlsxwriter") as writer:
    raw.to_excel(writer, sheet_name="Vendas_Originais", index=False)
    df_diario.reset_index().to_excel(writer, sheet_name="Receita_Diaria", index=False)
    forecast_df.to_excel(writer, sheet_name="Previsao_30_Dias", index=False)


fig_receita.write_image("plots/plot_receita_diaria.png")
fig_lucro.write_image("plots/plot_lucro_diario.png")
fig_lucro_capital.write_image("plots/plot_lucro_capital.png")
fig_forecast.write_image("plots/plot_previsao_30_dias.png")

wb = load_workbook(output_excel)
ws = wb.create_sheet("Dashboard")

ws["A1"] = "📊 Dashboard Katiana Store — Análise de Vendas, Lucro e Investimento"
ws["A1"].font = ws["A1"].font.copy(bold=True, size=16)

img1 = Image("plots/plot_receita_diaria.png")
img2 = Image("plots/plot_lucro_diario.png")
img3 = Image("plots/plot_lucro_capital.png")
img4 = Image("plots/plot_previsao_30_dias.png")

ws.add_image(img1, "A3")
ws.add_image(img2, "J3")
ws.add_image(img3, "A25")
ws.add_image(img4, "J25")

wb.save(output_excel)
print("✅ Dashboard completo com gráficos embutidos no Excel!")

print(f"\n🏁 Pipeline finalizado com sucesso!")
print(f"📘 Arquivo Excel: {output_excel}")
print(f"📊 Gráficos salvos em: {plots_path}")
